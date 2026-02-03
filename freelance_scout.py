import streamlit as st
import pandas as pd
from supabase import create_client, Client
import google.generativeai as genai
import requests
from io import BytesIO
from pypdf import PdfReader
from datetime import datetime
import json
import tempfile
import pathlib
from pydantic import BaseModel, Field
from typing import List, Optional
from streamlit_extras.add_vertical_space import add_vertical_space
from duckduckgo_search import DDGS
from firecrawl import FirecrawlApp
import xml.etree.ElementTree as ET
from streamlit_agraph import agraph, Node, Edge, Config
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# --- CONFIGURATION & MODELS ---
st.set_page_config(
    page_title="THE FACTORY",
    page_icon="🏭",
    layout="wide",
    initial_sidebar_state="expanded",
)

def generate_excel_matrix(papers):
    wb = Workbook()
    ws = wb.active
    ws.title = "Research Matrix"
    
    # Headers
    headers = ["Title", "Authors", "Year", "Status", "Impact Score", "Source", "Abstract/Summary"]
    ws.append(headers)
    
    # Styling Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="121212", end_color="121212", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # Data Rows
    for p in papers:
        authors = ", ".join(p.get('authors', [])) if isinstance(p.get('authors'), list) else "Anon"
        row = [
            p.get('title', 'Untitled'),
            authors,
            p.get('year', 'n.d.'),
            p.get('reading_status', 'Unread'),
            p.get('impact_score', 0.0),
            p.get('source_type', 'Unknown'),
            p.get('abstract', '')[:1000] # Truncate massive abstracts for Excel readability
        ]
        ws.append(row)
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        # Cap width
        if adjusted_width > 50: adjusted_width = 50
        ws.column_dimensions[column].width = adjusted_width
        
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

# Custom Premium CSS: THE FACTORY 2.0
st.markdown("""
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:ital,wght@0,400;0,700;1,400&family=Inter:wght@300;400;600&display=swap" rel="stylesheet">
<style>
    /* Absolute Ivory Force */
    .stApp, [data-testid="stAppViewContainer"], .main {
        background: #FAF9F6 !important;
        background-image: 
            radial-gradient(at 0% 0%, rgba(197, 160, 33, 0.03) 0px, transparent 50%),
            url("https://www.transparenttextures.com/patterns/natural-paper.png") !important;
        color: #121212 !important;
    }

    /* Total Header Deletion */
    header, [data-testid="stHeader"] {
        background: transparent !important;
        height: 0px !important;
        overflow: hidden !important;
        display: none !important;
    }
    [data-testid="stAppViewContainer"] {
        padding-top: 0px !important;
    }
    
    /* Force Sidebar Slim & Ink Contrast */
    [data-testid="stSidebar"] {
        width: 250px !important;
        background: #ffffff !important;
        border-right: 1px solid rgba(18, 18, 18, 0.05);
    }
    [data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
        padding: 1.5rem 1rem !important;
        gap: 0.2rem !important;
    }
    
    /* Navigation Visibility: INK ON IVORY */
    div[data-testid="stRadio"] label p {
        color: #121212 !important;
        font-weight: 500 !important;
        font-size: 0.8rem !important;
        opacity: 1 !important;
    }
    
    /* Removing Streamlit Radio Dots for Clean Menu */
    div[data-testid="stRadio"] input {
        display: none !important;
    }
    div[data-testid="stRadio"] div[role="radiogroup"] > label {
        padding: 8px 12px !important;
        margin-bottom: 4px !important;
        border-radius: 4px !important;
        transition: all 0.2s ease;
        background: transparent !important;
    }
    div[data-testid="stRadio"] div[role="radiogroup"] > label:hover {
        background: rgba(18, 18, 18, 0.03) !important;
    }
    div[data-testid="stRadio"] div[role="radiogroup"] [aria-checked="true"] {
        background: rgba(197, 160, 33, 0.08) !important;
        border-left: 3px solid #C5A021 !important;
    }

    /* Vertical Navigation Control Label */
    .nav-label {
        font-family: 'Inter', sans-serif;
        font-size: 0.65rem;
        text-transform: uppercase;
        letter-spacing: 0.25em;
        color: #C5A021;
        margin: 2rem 0 0.8rem 0;
        font-weight: 700;
    }

    /* Shrunk Boutique Buttons */
    .stButton>button {
        background: #121212 !important;
        color: #FAF9F6 !important;
        font-size: 0.65rem !important;
        font-weight: 700 !important;
        border-radius: 2px !important;
        padding: 0.3rem 0.6rem !important;
        width: 100% !important;
        text-transform: uppercase;
        letter-spacing: 0.1em;
        border: none !important;
        margin-bottom: 0px !important;
    }
    .stButton>button:hover {
        background: #C5A021 !important;
        box-shadow: 0 5px 15px rgba(197, 160, 33, 0.2);
    }
    
    /* Hero Section: The Factory Identity */
    .main-header {
        font-family: 'Playfair Display', serif !important;
        font-size: 4rem !important;
        font-weight: 700 !important;
        text-align: center;
        margin-top: 3rem !important;
        border-bottom: 2px solid #121212;
        padding-bottom: 1rem;
        animation: slideIn 1s cubic-bezier(0.23, 1, 0.32, 1);
    }
    @keyframes slideIn {
        from { opacity: 0; transform: translateY(-20px); }
        to { opacity: 1; transform: translateY(0); }
    }
    .sub-header {
        font-family: 'Inter', sans-serif;
        color: #C5A021;
        font-size: 0.75rem;
        text-transform: uppercase;
        letter-spacing: 0.5em;
        text-align: center;
        margin-top: 1rem;
        margin-bottom: 5rem;
    }

    /* Factory Cards: Boutique Feel */
    .factory-card {
        background: #ffffff;
        border: 1px solid rgba(18, 18, 18, 0.05);
        padding: 2.5rem;
        margin-bottom: 2rem;
        box-shadow: 0 5px 15px rgba(0,0,0,0.01);
    }

    /* Custom Selection Styling */
    .stSelectbox [data-baseweb="select"] {
        background: #ffffff !important;
        border: 1px solid rgba(18, 18, 18, 0.1) !important;
        font-size: 0.85rem !important;
    }
    
    /* High Contrast Text Inputs */
    .stTextInput input {
        background: #ffffff !important;
        color: #121212 !important;
        border: 1px solid rgba(18, 18, 18, 0.2) !important;
        padding: 0.8rem !important;
    }

    /* Research Cockpit Specifics */
    .intelligence-pane {
        background: #ffffff;
        border-right: 1px solid rgba(18, 18, 18, 0.1);
        padding-right: 1.5rem;
    }
    .synthesis-pane {
        background: #F4F4F2;
        padding-left: 1.5rem;
        height: 100vh;
        overflow-y: auto;
    }
    .feed-item {
        background: #ffffff;
        border: 1px solid rgba(18, 18, 18, 0.05);
        padding: 1rem;
        margin-bottom: 0.8rem;
        border-radius: 4px;
        transition: transform 0.2s ease;
    }
    .feed-item:hover {
        border-color: #C5A021;
        transform: translateX(3px);
    }
    .unified-bar {
        background: #121212;
        color: #FAF9F6;
        padding: 1rem;
        border-radius: 4px;
        margin-bottom: 2rem;
    }
    
    /* Scrollbar Ivory & Gold */
    ::-webkit-scrollbar { width: 5px; }
    ::-webkit-scrollbar-thumb { background: #C5A021; border-radius: 10px; }
    ::-webkit-scrollbar-track { background: transparent; }
</style>
""", unsafe_allow_html=True)

class PaperAnalysis(BaseModel):
    relevant: bool = Field(description="Is this paper highly relevant to the research query?")
    summary: str = Field(description="2-sentence academic summary focusing on findings.")
    methodology: str = Field(description="Primary research methodology used (e.g. Qualitative, Meta-analysis, Empirical).")

class Metadata(BaseModel):
    title: str = Field(description="Full academic title of the paper.")
    authors: List[str] = Field(description="List of all contributing authors.")
    year: Optional[int] = Field(description="Year of publication.")
    abstract: str = Field(description="Concise abstract or summary.")

class ScrapedContent(BaseModel):
    title: str = Field(description="The clear title of the article or paper.")
    author: str = Field(description="Name of the primary author or organization.")
    publication_date: str = Field(description="Date of publication (YYYY-MM-DD) or 'n.d.'")
    key_findings: List[str] = Field(description="List of 3-5 main takeaways or results.")
    methodology: str = Field(description="Brief description of how the information was gathered (e.g., 'Opinion', 'Case Study', 'Review').")
    summary: str = Field(description="A coherent paragraph summarizing the full content.")

class SynthesisResponse(BaseModel):
    answer: str
    citations: List[str]

@st.cache_resource
def init_supabase():
    return create_client(st.secrets["supabase"]["url"], st.secrets["supabase"]["key"])

@st.cache_resource
def init_gemini():
    genai.configure(api_key=st.secrets["google"]["api_key"])
    return genai.GenerativeModel('gemini-2.0-flash')

@st.cache_resource
def init_firecrawl():
    key = st.secrets["firecrawl"].get("api_key")
    return FirecrawlApp(api_key=key) if key else None

try:
    db = init_supabase()
    ai = init_gemini()
    firecrawl = init_firecrawl()
except Exception as e:
    st.error(f"Initialization Failed: {e}")
    st.stop()

def generate_word_xml_bib(papers):
    root = ET.Element("b:Sources", {
        "SelectedStyle": "\\APA.XSL",
        "xmlns:b": "http://schemas.openxmlformats.org/officeDocument/2006/bibliography",
        "xmlns": "http://schemas.openxmlformats.org/officeDocument/2006/bibliography"
    })
    for p in papers:
        source = ET.SubElement(root, "b:Source")
        ET.SubElement(source, "b:Tag").text = p['title'][:10] + str(p['year'])
        ET.SubElement(source, "b:SourceType").text = "JournalArticle"
        author_list = ET.SubElement(ET.SubElement(source, "b:Author"), "b:Author")
        names = ET.SubElement(author_list, "b:NameList")
        for a in p['authors']:
            person = ET.SubElement(names, "b:Person")
            ET.SubElement(person, "b:Last").text = a
        ET.SubElement(source, "b:Title").text = p['title']
        ET.SubElement(source, "b:Year").text = str(p['year'])
    return ET.tostring(root, encoding='utf-8')

# --- AUTHENTICATION ---
if "user" not in st.session_state:
    st.session_state.user = None
if "auth_mode" not in st.session_state:
    st.session_state.auth_mode = "Login" # Keep this line for initial state, though the new auth_gate might override it.

if st.session_state.user:
    try:
        db.postgrest.auth(st.session_state.user.access_token)
    except:
        pass

# Main Cinematic Header
st.markdown('<p class="sub-header" style="margin-top: 4rem;">EST. 2026</p>', unsafe_allow_html=True)
st.markdown('<h1 class="main-header">THE FACTORY</h1>', unsafe_allow_html=True)
st.markdown('<p class="sub-header">PRECISION RESEARCH ENGINE</p>', unsafe_allow_html=True)

# Auth Gate Logic with Factory Card
def auth_gate():
    if not st.session_state.user:
        st.markdown('<div class="factory-card" style="max-width:550px; margin: 100px auto auto;">', unsafe_allow_html=True)
        st.markdown('<h3 style="text-align:center; margin-bottom: 2rem;">Factory Access</h3>', unsafe_allow_html=True)
        
        auth_mode = st.radio("Access Level", ["Operator Login", "New Studio Enrollment"], horizontal=True)
        
        email = st.text_input("Personnel Email", placeholder="operator@factory.ai")
        password = st.text_input("Credentials", type="password", placeholder="••••••••")
        
        if auth_mode == "Operator Login":
            if st.button("Enter Factory"):
                try:
                    res = db.auth.sign_in_with_password({"email": email, "password": password})
                    st.session_state.user = res.user
                    st.rerun()
                except Exception as e:
                    st.error(f"Access Denied: {str(e)}")
        else:
            if st.button("Register Operator"):
                try:
                    res = db.auth.sign_up({"email": email, "password": password})
                    st.success("Enrollment requested. Please check your secure inbox.")
                except Exception as e:
                    st.error(f"Enrollment Error: {str(e)}")
        
        st.markdown('</div>', unsafe_allow_html=True)
        return False
    return True

if not auth_gate():
    st.stop()

# Sidebar: The Curator's Panel
with st.sidebar:
    st.markdown('<p class="sector-badge" style="text-align:left; margin-bottom:1rem;">THE CURATOR\'S PANEL</p>', unsafe_allow_html=True)
    st.markdown('<h2 style="font-size: 1.5rem; margin-top:0;">Terminal Control</h2>', unsafe_allow_html=True)
    
    with st.expander("👤 OPERATOR PROFILE", expanded=False):
        st.info(f"Identity: {st.session_state.user.email}")
        if st.button("Exit Factory"):
            db.auth.sign_out() # Changed supabase to db
            st.session_state.user = None
            st.rerun()

    st.markdown("---")
    # Project Selection / Allocation
    res = db.table("projects").select("*").eq("owner_id", st.session_state.user.id).execute()
    projects = res.data
    p_names = [p["name"] for p in projects]
    
    st.markdown("### Active Workspaces")
    # Determine active_project and project_id based on selection
    if projects:
        sel = st.selectbox("Select Workspace", p_names, key="sidebar_project_select")
        active_project = next(p for p in projects if p['name'] == sel)
        st.session_state.project_id = active_project['id']
    else:
        st.session_state.project_id = None
        st.info("No workspaces allocated.")
    
    st.markdown("<br>", unsafe_allow_html=True)
    with st.popover("➕ NEW ALLOCATION", use_container_width=True):
        new_p = st.text_input("New Workspace Designation")
        if st.button("INITIATE WORKSPACE") and new_p:
            db.table("projects").insert({
                "name": new_p, 
                "client_name": "Agency",
                "owner_id": st.session_state.user.id
            }).execute()
            st.rerun()

    if st.session_state.project_id:
        st.markdown("<br>", unsafe_allow_html=True)
        with st.popover("🗑️ DELETE", use_container_width=True):
            st.warning("This will permanently incinerate all intelligence records within this workspace.")
            if st.button("CONFIRM DELETE"):
                try:
                    db.table("papers").delete().eq("project_id", st.session_state.project_id).execute()
                    db.table("projects").delete().eq("id", st.session_state.project_id).execute()
                    st.session_state.project_id = None
                    st.success("Workspace Deleted.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Deletion Failed: {e}")

    st.markdown('<p class="nav-label">TERMINAL STATUS</p>', unsafe_allow_html=True)
    st.success("COCKPIT ACTIVE")
    st.markdown("---")
    st.markdown("<p style='font-size: 0.6rem; color: #121212; text-align: center; opacity: 0.5;'>THE FACTORY<br>© 2026 ANTIGRAVITY AI</p>", unsafe_allow_html=True)

# --- MAIN UI ---
if st.session_state.project_id:
    # Top Sector Status
    st.markdown(f'<div class="sector-badge">WORKSPACE: {active_project["name"].upper()} | ID: {active_project["id"][:8].upper()}</div>', unsafe_allow_html=True)
    
    # Split Screen: THE RESEARCH COCKPIT
    intel_col, synth_col = st.columns([1.2, 1], gap="large")

    # --- LEFT PANE: INTELLIGENCE FEED ---
    with intel_col:
        st.markdown("### 🛰️ Intelligence Feed")
        
        # Unified Command Bar
        with st.container():
            st.markdown('<div class="unified-bar">', unsafe_allow_html=True)
            
            mode = st.radio("Mission Mode", ["Search", "Ingest PDF", "Scrape URL"], horizontal=True)
            
import tempfile
import pathlib

# ... (rest of imports)

# ... inside Main UI ...

            if mode == "Ingest PDF":
                up = st.file_uploader("Source PDF", type="pdf", label_visibility="collapsed")
                if st.button("EXECUTE INGEST", use_container_width=True) and up:
                    with st.spinner("Uploading & Analyzing (Vision Engine)..."):
                        # 1. Save to Temp File
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
                            tmp.write(up.getvalue())
                            tmp_path = tmp.name

                        try:
                            # 2. Upload to Gemini
                            pdf_file = genai.upload_file(tmp_path, mime_type="application/pdf")
                            
                            # 3. Two-Pass Extraction (Vision)
                            # Pass A: Strict Metadata
                            meta_res = ai.generate_content(
                                [pdf_file, "Extract strict academic metadata."], 
                                generation_config={"response_mime_type": "application/json", "response_schema": ScrapedContent}
                            ).text
                            meta = ScrapedContent.model_validate_json(meta_res)
                            
                            # Pass B: Full Content Ingest
                            content_res = ai.generate_content([pdf_file, "Transcribe this full document into clean Markdown. Describe all charts/images."]).text

                            # 4. Archive
                            rich_abstract = f"""{meta.summary}
                            
                            KEY FINDINGS:
                            {chr(10).join(['- ' + k for k in meta.key_findings])}
                            
                            METHODOLOGY: {meta.methodology}
                            """

                            db.table("papers").insert({
                                "project_id": st.session_state.project_id,
                                "title": meta.title,
                                "authors": [meta.author], # Schema expects list
                                "year": int(meta.publication_date[:4]) if meta.publication_date[0].isdigit() else datetime.now().year,
                                "abstract": rich_abstract,
                                "content_body": content_res, # New Full Text Column
                                "source_type": "pdf"
                            }).execute()
                            st.toast("PDF Digitized & Archived (Vision-Enhanced)!")
                            
                        except Exception as e:
                            st.error(f"Ingest Error: {e}")
                        finally:
                            pathlib.Path(tmp_path).unlink(missing_ok=True)
            else:
                cmd = st.text_input("UNIFIED COMMAND", key="unified_cmd", placeholder="Enter keywords or URL...", label_visibility="collapsed")
                c1, c2 = st.columns(2)
                grey_lit = c1.toggle("Grey Lit", help="Search beyond academic journals")
                auto_pilot = c2.toggle("Auto-Pilot", help="Autonomous keyword refinement")
                
                # Initialize Session State for Auto-Pilot
                if "pilot_proposal" not in st.session_state: st.session_state.pilot_proposal = None

                if st.button("EXECUTE MISSION", use_container_width=True):
                    try:
                        if mode == "Search" and cmd:
                            with st.status("Executing Scout...") as status:
                                def perform_search(term):
                                    try:
                                        ss = requests.get(f"https://api.semanticscholar.org/graph/v1/paper/search?query={term}&limit=10&fields=title,authors,year,abstract,url", timeout=10).json().get("data", [])
                                    except Exception:
                                        ss = []
                                    
                                    if grey_lit:
                                        try:
                                            with DDGS() as ddgs:
                                                for dr in ddgs.text(f"{term} filetype:pdf", max_results=5):
                                                    ss.append({"title": dr['title'], "url": dr['href'], "abstract": dr['body'], "authors": [{"name": "Web Source"}], "year": datetime.now().year, "source_type": "grey"})
                                        except Exception as e:
                                            st.write(f"Grey Lit Warning: {e}")
                                    return ss
                                
                                results = perform_search(cmd)
                                st.session_state.current_results = results
                                
                                # Auto-Pilot Trigger
                                if auto_pilot and len(results) < 10:
                                    with st.spinner("Auto-Pilot: Analyzing Research Void..."):
                                        strategy_prompt = f"""
                                        The user searched for '{cmd}' but found few results. 
                                        Act as a Senior Research Librarian. 
                                        1. Analyze WHY (too specific? wrong terminology?).
                                        2. Suggest 3 alternative academic search queries (broadening, synonyms, or methodology-specific).
                                        3. Provide a brief rationale.
                                        Return JSON: {{ "analysis": "...", "queries": ["q1", "q2", "q3"] }}
                                        """
                                        try:
                                            strategy = json.loads(ai.generate_content(strategy_prompt, generation_config={"response_mime_type": "application/json"}).text)
                                            st.session_state.pilot_proposal = strategy
                                        except Exception as e:
                                            st.warning(f"Auto-Pilot Glitch: {e}")
                                else:
                                    st.toast("Scout Mission Complete.")
                                status.update(label="Scout Complete", state="complete")

                        elif mode == "Scrape URL" and cmd and firecrawl:
                             with st.spinner("Scraping Sector..."):
                                try:
                                    sc = firecrawl.scrape_url(cmd, params={'formats': ['markdown']})
                                    md = sc.get('markdown','')
                                    
                                    if not md: raise ValueError("No markdown content returned.")

                                    # Validated Extraction
                                    extract_prompt = f"Analyze this web content and extract strict metadata: {md[:8000]}"
                                    ai_res_json = ai.generate_content(
                                        extract_prompt, 
                                        generation_config={"response_mime_type": "application/json", "response_schema": ScrapedContent}
                                    ).text
                                    validated_data = ScrapedContent.model_validate_json(ai_res_json)
                                    
                                    # formatted abstract with deep metadata
                                    rich_abstract = f"""{validated_data.summary}
                                    
                                    KEY FINDINGS:
                                    {chr(10).join(['- ' + k for k in validated_data.key_findings])}
                                    
                                    METHODOLOGY: {validated_data.methodology}
                                    """
                                    
                                    db.table("papers").insert({
                                        "project_id": st.session_state.project_id, 
                                        "title": validated_data.title,
                                        "authors": [validated_data.author],
                                        "year": int(validated_data.publication_date[:4]) if validated_data.publication_date[0].isdigit() else datetime.now().year,
                                        "abstract": rich_abstract, 
                                        "url": cmd, 
                                        "source_type": "web"
                                    }).execute()
                                    st.toast("Web Intelligence Captured & Validated!")
                                except Exception as e:
                                    st.error(f"Scrape Mission Failed: {e}")
                    except Exception as critical_e:
                        st.error(f"Terminal Error: {critical_e}")

                # Auto-Pilot Proposal UI
                if st.session_state.pilot_proposal:
                    with st.container():
                        st.markdown(f"""
                        <div style="background: #FFFBF0; border: 1px solid #C5A021; padding: 1rem; border-radius: 4px; margin-top: 1rem;">
                            <h4 style="color:#C5A021; margin:0;">✈️ Auto-Pilot Proposal</h4>
                            <p style="font-size:0.8rem;"><b>Analysis:</b> {st.session_state.pilot_proposal['analysis']}</p>
                            <p style="font-size:0.8rem; font-weight:bold;">Proposed Trajectories:</p>
                        </div>
                        """, unsafe_allow_html=True)
                        
                        cols = st.columns(3)
                        for i, q in enumerate(st.session_state.pilot_proposal['queries']):
                            if cols[i].button(q, key=f"pilot_{i}", use_container_width=True):
                                new_res = perform_search(q)
                                st.session_state.current_results += new_res
                                st.session_state.pilot_proposal = None # Clear after execution
                                st.rerun()

            st.markdown('</div>', unsafe_allow_html=True)

        # Feed Stream
        feed_tabs = st.tabs(["Review Queue", "Research Archive", "Network Graph", "Handover"])
        
        with feed_tabs[0]: # Review Queue (Current Results)
            if "current_results" in st.session_state:
                for p in st.session_state.current_results:
                    with st.container():
                        st.markdown(f"""
                        <div class="feed-item">
                            <p style="font-size:0.7rem; color:#C5A021; margin:0;">{p.get('source_type', 'ACADEMIC').upper()}</p>
                            <h4 style="margin:0.2rem 0;">{p['title']}</h4>
                            <p style="font-size:0.8rem; opacity:0.7;">{p.get('year', 'N/A')} | {', '.join([a.get('name', '') for a in p.get('authors', [])]) if isinstance(p.get('authors'), list) else 'Anon'}</p>
                        </div>
                        """, unsafe_allow_html=True)
                        if st.button("Save to Library", key=f"save_{p['title']}"):
                            db.table("papers").insert({
                                "project_id": st.session_state.project_id,
                                "title": p['title'], "authors": [a.get('name','Anon') for a in p.get('authors',[])] if isinstance(p.get('authors'), list) else ['Anon'],
                                "year": p.get('year'), "abstract": p.get('abstract'), "url": p.get('url'), "source_type": p.get('source_type', 'scout')
                            }).execute()
                            st.toast("Archived!")
            else:
                st.info("Feed is silent. Initiate a mission above.")

        with feed_tabs[1]: # Research Archive (Citation Matrix)
            paps = db.table("papers").select("*").eq("project_id", st.session_state.project_id).order("created_at", desc=True).execute().data or []
            if paps:
                df = pd.DataFrame(paps)
                
                # Ensure columns exist for the editor
                if 'reading_status' not in df.columns: df['reading_status'] = 'Unread'
                if 'tags' not in df.columns: df['tags'] = [[] for _ in range(len(df))]
                if 'citation_count' not in df.columns: df['citation_count'] = 0

                st.markdown("### 📚 Citation Matrix")
                
                edited_df = st.data_editor(
                    df,
                    column_config={
                        "title": st.column_config.TextColumn("Title", disabled=True, width="medium"),
                        "reading_status": st.column_config.SelectboxColumn(
                            "Status",
                            options=["Unread", "Reading", "Synthesized"],
                            required=True
                        ),
                        "tags": st.column_config.ListColumn("Tags"),
                        "citation_count": st.column_config.NumberColumn("Citations", disabled=True),
                        "year": st.column_config.NumberColumn("Year", disabled=True),
                        "source_type": st.column_config.TextColumn("Source", disabled=True)
                    },
                    column_order=["title", "reading_status", "tags", "year", "citation_count", "source_type"],
                    hide_index=True,
                    use_container_width=True,
                    key="citation_matrix"
                )

                # Commit Changes
                if st.session_state.citation_matrix.get("edited_rows"):
                    changes = st.session_state.citation_matrix["edited_rows"]
                    with st.spinner("Syncing Matrix..."):
                        for idx, diff in changes.items():
                            paper_id = df.iloc[idx]["id"]
                            db.table("papers").update(diff).eq("id", paper_id).execute()
                    st.toast("Matrix Synced.")
                    st.rerun()

                st.markdown("---")
                
                # Contextual Actions for Selected Paper (Mock selection via selectbox for now as data_editor selection is beta)
                sel_paper = st.selectbox("Select Paper for Deep Actions", df['title'].tolist())
                sel_p_data = df[df['title'] == sel_paper].iloc[0]

                c1, c2 = st.columns(2)
                with c1:
                    if st.button("⛏️ Snowball Mine", key="mine_btn"):
                         with st.spinner("Mining Research Graph..."):
                            ss_search = requests.get(f"https://api.semanticscholar.org/graph/v1/paper/search?query={sel_p_data['title']}&limit=1&fields=paperId").json()
                            if ss_search.get("data"):
                                pid = ss_search["data"][0]["paperId"]
                                ref_url = f"https://api.semanticscholar.org/graph/v1/paper/{pid}/references?limit=10&fields=title,authors,year,abstract,url,citationCount"
                                refs = requests.get(ref_url).json().get("data", [])
                                count = 0
                                for r_item in refs:
                                    r = r_item.get("citedPaper")
                                    if r and r.get("title"):
                                        # Check duplicate
                                        exists = db.table("papers").select("id").eq("title", r['title']).eq("project_id", st.session_state.project_id).execute().data
                                        if not exists:
                                            db.table("papers").insert({
                                                "project_id": st.session_state.project_id,
                                                "title": r['title'], "authors": [a.get('name', 'Unknown') for a in r.get('authors', [])],
                                                "year": r.get('year'), "abstract": r.get('abstract'), "url": r.get('url'),
                                                "citation_count": r.get('citationCount', 0),
                                                "source_type": "snowball-mining"
                                            }).execute()
                                            count += 1
                                st.success(f"Mined {count} new references!")
                                st.rerun()
                            else: st.error("Paper not found in Graph.")
                with c2:
                    st.write(f"**Abstract:** {sel_p_data.get('abstract', '')[:300]}...")

            else:
                st.info("Archive empty.")

        with feed_tabs[2]: # Network Graph
            paps = db.table("papers").select("*").eq("project_id", st.session_state.project_id).execute().data or []
            if paps:
                nodes = []
                edges = []
                # Create Nodes
                for p in paps:
                    # Size based on citation count (logarithmic scale approx)
                    size = 15 + (p.get('citation_count', 0) / 10) 
                    nodes.append(Node(id=p['id'], label=p['title'][:20]+"...", size=min(size, 40), shape="dot", color="#C5A021"))
                
                # Create Edges (Shared Authors)
                for i, p1 in enumerate(paps):
                    for p2 in paps[i+1:]:
                        # Find intersection of authors
                        authors1 = set(p1.get('authors', []))
                        authors2 = set(p2.get('authors', []))
                        shared = authors1.intersection(authors2)
                        if shared:
                            edges.append(Edge(source=p1['id'], target=p2['id'], label=list(shared)[0], color="#121212"))

                config = Config(width="100%", height=500, directed=False, physics=True, hierarchical=False)
                return_value = agraph(nodes=nodes, edges=edges, config=config)
            else:
                st.info("Add papers to visualize the network.")

        with feed_tabs[3]: # Handover Dossier
            p_exp = db.table("papers").select("*").eq("project_id", st.session_state.project_id).execute().data or []
            if p_exp:
                st.download_button("Word Bib (XML)", generate_word_xml_bib(p_exp), "bib.xml", use_container_width=True)
                
                # Excel Export using new function
                excel_data = generate_excel_matrix(p_exp)
                st.download_button("Excel Research Matrix (Pro)", excel_data, "matrix.xlsx", use_container_width=True)
                
                if st.button("🚀 Sync to Notion", use_container_width=True):
                    token = st.secrets["notion"].get("api_token")
                    db_id = st.secrets["notion"].get("database_id")
                    if token and db_id:
                        with st.status("Syncing to Notion Workspace...") as status:
                            headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json", "Notion-Version": "2022-06-28"}
                            progress_bar = st.progress(0)
                            success_count = 0
                            error_count = 0
                            
                            for i, p in enumerate(p_exp):
                                try:
                                    data = {
                                        "parent": {"database_id": db_id}, 
                                        "properties": {
                                            "Title": {"title": [{"text": {"content": p['title'][:2000]}}]}, # Limit title length
                                            "Year": {"number": p['year']} if p['year'] else None,
                                            "Authors": {"rich_text": [{"text": {"content": ", ".join(p.get('authors', []))[:2000]}}]},
                                            "URL": {"url": p['url']} if p['url'] else None,
                                            "Status": {"select": {"name": p.get('reading_status', 'Unread')}} # Sync Status
                                        }
                                    }
                                    # Clean None values
                                    data["properties"] = {k: v for k, v in data["properties"].items() if v is not None}
                                    
                                    res = requests.post("https://api.notion.com/v1/pages", headers=headers, json=data)
                                    res.raise_for_status()
                                    success_count += 1
                                except Exception as e:
                                    error_count += 1
                                    st.write(f"⚠️ Failed to sync '{p['title'][:20]}...': {e}")
                                
                                progress_bar.progress((i + 1) / len(p_exp))
                            
                            status.update(label=f"Notion Sync Complete: {success_count} Synced, {error_count} Failed", state="complete")
                            st.toast(f"Sync Complete: {success_count} Success, {error_count} Failures")
                    else: st.warning("Notion credentials missing in secrets.")

    # --- RIGHT PANE: SYNTHESIS WORKSPACE ---
    with synth_col:
        st.markdown("### ✍️ Synthesis Workspace")
        
        # Pull context from library
        papers_rag = db.table("papers").select("*").eq("project_id", st.session_state.project_id).execute().data or []
        
        # Construct High-Fidelity Context
        kb_context = ""
        for p in papers_rag:
            auth_str = p['authors'][0] if p.get('authors') and len(p['authors']) > 0 else 'Anon'
            year_str = str(p.get('year', 'n.d.'))
            # Fallback to abstract if content_body is empty
            content = p.get('content_body', '') 
            if not content: content = p.get('abstract', '')
            
            kb_context += f"""
            ---
            REF_CODE: [{auth_str}, {year_str}]
            TITLE: {p['title']}
            CONTENT: {content[:3000]} 
            ---
            """
        
        if "messages" not in st.session_state:
            st.session_state.messages = []
            
        for msg in st.session_state.messages:
            with st.chat_message(msg["role"]):
                st.markdown(msg["content"])
                
        if prompt := st.chat_input("Synthesize intelligence or draft report..."):
            st.session_state.messages.append({"role": "user", "content": prompt})
            with st.chat_message("user"):
                st.markdown(prompt)
                
            with st.chat_message("assistant"):
                with st.spinner("Synthesizing (Strict Grounding)..."):
                    system_prompt = """
                    You are an Elite Research Assistant. Your goal is to synthesize answers using ONLY the provided Context Library.
                    
                    RULES:
                    1. CITATIONS MANDATORY: Every single claim or fact must be immediately followed by its source in [Author, Year] format.
                    2. NO HALLUCINATION: If the answer is not in the Context Library, state clearly: "This information is not present in the current archives."
                    3. ACADEMIC TONE: Maintain a professional, objective, and analytical tone.
                    4. SYNTHESIS: Do not just list summaries. Connect the dots between papers. Compare and contrast findings.
                    """
                    
                    full_prompt = f"{system_prompt}\n\nCONTEXT LIBRARY:\n{kb_context}\n\nUSER PROMPT: {prompt}"
                    
                    response = ai.generate_content(full_prompt).text
                    st.markdown(response)
                    st.session_state.messages.append({"role": "assistant", "content": response})

else:
    st.info("Unlock a Workspace via Sidebar to Activate Terminal.")
